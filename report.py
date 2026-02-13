"""Excel report generation for attendance data."""
import calendar
import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import db


async def generate_attendance_report(teacher_id: int, year: int, month: int) -> io.BytesIO:
    """Generate an Excel attendance report for a teacher's class for a given month.

    Returns a BytesIO buffer containing the .xlsx file.
    """
    all_teachers = await db.get_all_teachers()
    teacher = next((t for t in all_teachers if t["id"] == teacher_id), None)
    teacher_name = teacher["name"] if teacher else "غير معروف"

    students = await db.get_students_by_teacher(teacher_id)
    attendance_records = await db.get_attendance_for_month(teacher_id, year, month)
    attendance_dates = await db.get_attendance_dates_for_month(teacher_id, year, month)

    # Build a set of (student_id, date_str) for quick lookup
    attendance_set: set[tuple[int, str]] = set()
    for record in attendance_records:
        if record["date"]:
            attendance_set.add((record["student_id"], record["date"]))

    month_name = calendar.month_name[month]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=11)
    cell_font = Font(size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    # ── Title rows ────────────────────────────────────────────────────────
    # Column count is dynamic: Student Name + N dates + Total
    date_col_count = len(attendance_dates)
    last_col = 1 + date_col_count + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value=f"تقرير الحضور — {month_name} {year}")
    title_cell.font = header_font
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    teacher_cell = ws.cell(row=2, column=1, value=f"المعلم: {teacher_name}")
    teacher_cell.font = sub_header_font
    teacher_cell.alignment = Alignment(horizontal="center")

    # ── Header row ────────────────────────────────────────────────────────
    header_row = 4
    name_header = ws.cell(row=header_row, column=1, value="اسم الطالب")
    name_header.font = header_text
    name_header.fill = header_fill
    name_header.border = thin_border
    name_header.alignment = center

    for idx, date_str in enumerate(attendance_dates):
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        label = dt.strftime("%d-%B-%Y")
        cell = ws.cell(row=header_row, column=idx + 2, value=label)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    total_header = ws.cell(row=header_row, column=last_col, value="المجموع")
    total_header.font = header_text
    total_header.fill = header_fill
    total_header.border = thin_border
    total_header.alignment = center

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, student in enumerate(students):
        row = header_row + 1 + i
        name_cell = ws.cell(row=row, column=1, value=student["name"])
        name_cell.font = cell_font
        name_cell.border = thin_border

        total = 0
        for idx, date_str in enumerate(attendance_dates):
            cell = ws.cell(row=row, column=idx + 2)
            cell.border = thin_border
            cell.alignment = center
            cell.font = cell_font
            if (student["id"], date_str) in attendance_set:
                cell.value = "✓"
                total += 1

        total_cell = ws.cell(row=row, column=last_col, value=total)
        total_cell.font = Font(bold=True, size=10)
        total_cell.border = thin_border
        total_cell.alignment = center

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 25
    for idx in range(date_col_count):
        col_letter = get_column_letter(idx + 2)
        ws.column_dimensions[col_letter].width = 16
    total_col_letter = get_column_letter(last_col)
    ws.column_dimensions[total_col_letter].width = 7

    # ── Save to buffer ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
